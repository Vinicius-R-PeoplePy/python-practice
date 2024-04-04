#This code fills 289 cells of an Excel file i'm working on. 
#It searches for the matching terms and return the data as the matching calls for filling data in a certain cell. 

from openpyxl import load_workbook

# Load the Excel file
workbook = load_workbook('python/vlookup ovni(1).xlsx')

# Get the sheets
students_sheet = workbook['Students']
fees_sheet = workbook['Fees']

# Prepare the VLOOKUP formula
vlookup_formula = "=VLOOKUP(Fees!$B$4, Fees!$B$1:$B$24, 1, FALSE)"

# Iterate over each cell in column H of 'Students' sheet
for row in range(5, 294):  # Assuming data starts from row 5
    program_name = students_sheet.cell(row=row, column=7).value  # Assuming Major is in column G
    if program_name is None:
        continue  # Skip empty cells
    # Find the corresponding row in the 'Fees' sheet
    for fees_row in range(1, 25):  # Range $A$1:A$24
        if fees_sheet.cell(row=fees_row, column=1).value == program_name:  # Assuming Program is in column A
            # Write VLOOKUP formula to the corresponding cell in column H of 'Students' sheet
            students_sheet.cell(row=row, column=8).value = f"=VLOOKUP(Fees!B{fees_row}, Fees!$B$1:$B$24, 1, FALSE)"
            break

# Save the changes
workbook.save('python/vlookupautomated.xlsx')
